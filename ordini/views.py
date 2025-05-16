from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView,
    TemplateView, FormView, View
)
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin, PermissionRequiredMixin
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse, Http404, HttpResponseRedirect
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.db.models import Q, Sum, Count, Avg, F, Prefetch
from django.utils import timezone
from datetime import date, timedelta
from decimal import Decimal
import json
import csv
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.template.loader import render_to_string
from django.core.mail import send_mail
from django.conf import settings

from .models import (
    Categoria, Prodotto, Ordine, Ricezione, ProdottoRicevuto, Magazzino
)
from .forms import (
    CategoriaForm, ProdottoForm, OrdineForm, AggiornaStatoOrdineForm,
    RicezioneForm, ProdottoRicevutoForm, OrdineSearchForm, MagazzinoForm,
    MovimentoMagazzinoForm, MagazzinoFilterForm, ExportOrdiniForm,
    ReportOrdiniForm, QuickOrderForm, BulkActionForm, ProdottoRicevutoFormSet
)
from anagrafica.models import Fornitore
from dipendenti.models import Dipendente


# ====================== MIXINS ======================

class StaffRequiredMixin(UserPassesTestMixin):
    """Mixin per verificare che l'utente sia staff"""
    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser

    def handle_no_permission(self):
        messages.error(self.request, "Non hai i permessi per accedere a questa sezione")
        return redirect('home:index')


class OrdineAccessMixin(UserPassesTestMixin):
    """Mixin per verificare che l'utente possa accedere all'ordine"""
    def test_func(self):
        if self.request.user.is_staff or self.request.user.is_superuser:
            return True
        # Altri controlli di accesso se necessario
        return True


# ====================== DASHBOARD ======================

class DashboardOrdiniView(LoginRequiredMixin, TemplateView):
    """Dashboard principale ordini con statistiche"""
    template_name = 'ordini/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Statistiche generali
        context['totale_ordini'] = Ordine.objects.count()
        context['ordini_bozza'] = Ordine.objects.bozze().count()
        context['ordini_da_ricevere'] = Ordine.objects.da_ricevere().count()
        context['ordini_in_ritardo'] = Ordine.objects.in_ritardo().count()
        
        # Prodotti con scorte basse
        context['prodotti_scorte_basse'] = Magazzino.objects.scorte_basse().count()
        context['prodotti_in_scadenza'] = Magazzino.objects.in_scadenza().count()
        
        # Ordini recenti (ultimi 10)
        context['ordini_recenti'] = Ordine.objects.select_related(
            'prodotto', 'fornitore'
        )[:10]
        
        # Ultimi prodotti ricevuti
        context['prodotti_ricevuti_recenti'] = ProdottoRicevuto.objects.select_related(
            'prodotto', 'ricezione__ordine'
        ).order_by('-ricezione__data_ricezione')[:5]
        
        # Top fornitori (con più ordini)
        context['top_fornitori'] = Fornitore.objects.annotate(
            num_ordini=Count('ordine')
        ).filter(num_ordini__gt=0).order_by('-num_ordini')[:5]
        
        # Statistiche finanziarie (ultimi 30 giorni)
        data_limite = date.today() - timedelta(days=30)
        ordini_recenti = Ordine.objects.filter(
            data_creazione_ordine__gte=data_limite,
            status__in=[Ordine.StatusOrdine.RICEVUTO, Ordine.StatusOrdine.COMPLETATO]
        )
        context['totale_speso_30_giorni'] = ordini_recenti.aggregate(
            totale=Sum('totale_ordine_ivato')
        )['totale'] or Decimal('0.00')
        
        return context


# ====================== CATEGORIE ======================

class CategoriaListView(LoginRequiredMixin, ListView):
    """Lista delle categorie"""
    model = Categoria
    template_name = 'ordini/categorie/elenco.html'
    context_object_name = 'categorie'
    paginate_by = 20

    def get_queryset(self):
        queryset = Categoria.objects.all()
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(nome_categoria__icontains=q) |
                Q(descrizione__icontains=q)
            )
        return queryset.order_by('ordinamento', 'nome_categoria')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('q', '')
        return context


class CategoriaDetailView(LoginRequiredMixin, DetailView):
    """Dettaglio categoria con prodotti associati"""
    model = Categoria
    template_name = 'ordini/categorie/dettaglio.html'
    context_object_name = 'categoria'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Prodotti della categoria
        context['prodotti'] = self.object.prodotto_set.filter(attivo=True)
        context['numero_prodotti'] = context['prodotti'].count()
        return context


class CategoriaCreateView(LoginRequiredMixin, StaffRequiredMixin, CreateView):
    """Creazione nuova categoria"""
    model = Categoria
    form_class = CategoriaForm
    template_name = 'ordini/categorie/nuovo.html'
    success_url = reverse_lazy('ordini:elenco_categorie')

    def form_valid(self, form):
        messages.success(self.request, f'Categoria "{form.instance.nome_categoria}" creata con successo!')
        return super().form_valid(form)


class CategoriaUpdateView(LoginRequiredMixin, StaffRequiredMixin, UpdateView):
    """Modifica categoria esistente"""
    model = Categoria
    form_class = CategoriaForm
    template_name = 'ordini/categorie/modifica.html'

    def get_success_url(self):
        return reverse_lazy('ordini:dettaglio_categoria', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, f'Categoria "{form.instance.nome_categoria}" aggiornata con successo!')
        return super().form_valid(form)


class CategoriaDeleteView(LoginRequiredMixin, StaffRequiredMixin, DeleteView):
    """Eliminazione categoria"""
    model = Categoria
    template_name = 'ordini/categorie/elimina.html'
    success_url = reverse_lazy('ordini:elenco_categorie')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Verifica se la categoria ha prodotti associati
        context['prodotti_associati'] = self.object.prodotto_set.count()
        return context

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        # Verifica se ci sono prodotti associati
        if self.object.prodotto_set.exists():
            messages.error(request, "Impossibile eliminare: categoria ha prodotti associati")
            return redirect('ordini:dettaglio_categoria', pk=self.object.pk)
        
        success_url = self.get_success_url()
        nome_categoria = self.object.nome_categoria
        self.object.delete()
        messages.success(request, f'Categoria "{nome_categoria}" eliminata con successo!')
        return HttpResponseRedirect(success_url)


# ====================== PRODOTTI ======================

class ProdottoListView(LoginRequiredMixin, ListView):
    """Lista dei prodotti con filtri avanzati"""
    model = Prodotto
    template_name = 'ordini/prodotti/elenco.html'
    context_object_name = 'prodotti'
    paginate_by = 20

    def get_queryset(self):
        queryset = Prodotto.objects.select_related('categoria')
        
        # Filtro ricerca
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(nome_prodotto__icontains=q) |
                Q(ean__icontains=q) |
                Q(codice_interno__icontains=q) |
                Q(descrizione__icontains=q)
            )
        
        # Filtro categoria
        categoria_id = self.request.GET.get('categoria')
        if categoria_id:
            queryset = queryset.filter(categoria_id=categoria_id)
        
        # Filtro stato attivo
        attivo = self.request.GET.get('attivo')
        if attivo == 'true':
            queryset = queryset.filter(attivo=True)
        elif attivo == 'false':
            queryset = queryset.filter(attivo=False)
        
        # Filtro scorte basse
        scorte_basse = self.request.GET.get('scorte_basse')
        if scorte_basse == 'true':
            queryset = queryset.annotate(
                quantita_totale=Sum('magazzino__quantita_in_magazzino')
            ).filter(
                Q(quantita_totale__lte=F('scorta_minima')) |
                Q(quantita_totale__isnull=True)
            )
        
        return queryset.order_by('categoria__nome_categoria', 'nome_prodotto')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categorie'] = Categoria.objects.filter(attiva=True)
        context['search_query'] = self.request.GET.get('q', '')
        context['categoria_selezionata'] = self.request.GET.get('categoria', '')
        context['attivo_filtro'] = self.request.GET.get('attivo', '')
        context['scorte_basse_filtro'] = self.request.GET.get('scorte_basse', '')
        return context


class ProdottoDetailView(LoginRequiredMixin, DetailView):
    """Dettaglio prodotto con informazioni magazzino e ordini"""
    model = Prodotto
    template_name = 'ordini/prodotti/dettaglio.html'
    context_object_name = 'prodotto'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Informazioni magazzino
        magazzino_entries = Magazzino.objects.filter(
            prodotto=self.object
        ).order_by('data_scadenza')
        context['magazzino_entries'] = magazzino_entries
        context['quantita_totale'] = magazzino_entries.aggregate(
            totale=Sum('quantita_in_magazzino')
        )['totale'] or 0
        
        # Ordini recenti del prodotto
        context['ordini_recenti'] = Ordine.objects.filter(
            prodotto=self.object
        ).select_related('fornitore').order_by('-data_creazione_ordine')[:10]
        
        # Statistiche ordini
        context['totale_ordinato'] = Ordine.objects.filter(
            prodotto=self.object,
            status__in=[Ordine.StatusOrdine.RICEVUTO, Ordine.StatusOrdine.COMPLETATO]
        ).aggregate(totale=Sum('quantita_ordinata'))['totale'] or 0
        
        # Prezzo medio dell'ultimo ordine
        ultimo_ordine = Ordine.objects.filter(
            prodotto=self.object,
            status__in=[Ordine.StatusOrdine.RICEVUTO, Ordine.StatusOrdine.COMPLETATO]
        ).order_by('-data_creazione_ordine').first()
        context['ultimo_prezzo'] = ultimo_ordine.prezzo_unitario_ordine if ultimo_ordine else None
        
        return context


class ProdottoCreateView(LoginRequiredMixin, StaffRequiredMixin, CreateView):
    """Creazione nuovo prodotto"""
    model = Prodotto
    form_class = ProdottoForm
    template_name = 'ordini/prodotti/nuovo.html'
    success_url = reverse_lazy('ordini:elenco_prodotti')

    def form_valid(self, form):
        messages.success(self.request, f'Prodotto "{form.instance.nome_prodotto}" creato con successo!')
        return super().form_valid(form)


class ProdottoUpdateView(LoginRequiredMixin, StaffRequiredMixin, UpdateView):
    """Modifica prodotto esistente"""
    model = Prodotto
    form_class = ProdottoForm
    template_name = 'ordini/prodotti/modifica.html'

    def get_success_url(self):
        return reverse_lazy('ordini:dettaglio_prodotto', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, f'Prodotto "{form.instance.nome_prodotto}" aggiornato con successo!')
        return super().form_valid(form)


class ProdottoDeleteView(LoginRequiredMixin, StaffRequiredMixin, DeleteView):
    """Eliminazione prodotto"""
    model = Prodotto
    template_name = 'ordini/prodotti/elimina.html'
    success_url = reverse_lazy('ordini:elenco_prodotti')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Verifica se il prodotto ha ordini associati
        context['ordini_associati'] = self.object.ordine_set.count()
        context['magazzino_presente'] = Magazzino.objects.filter(
            prodotto=self.object,
            quantita_in_magazzino__gt=0
        ).exists()
        return context

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        # Verifica se ci sono ordini o magazzino
        if self.object.ordine_set.exists():
            messages.error(request, "Impossibile eliminare: prodotto ha ordini associati")
            return redirect('ordini:dettaglio_prodotto', pk=self.object.pk)
        
        if Magazzino.objects.filter(prodotto=self.object, quantita_in_magazzino__gt=0).exists():
            messages.error(request, "Impossibile eliminare: prodotto presente in magazzino")
            return redirect('ordini:dettaglio_prodotto', pk=self.object.pk)
        
        success_url = self.get_success_url()
        nome_prodotto = self.object.nome_prodotto
        self.object.delete()
        messages.success(request, f'Prodotto "{nome_prodotto}" eliminato con successo!')
        return HttpResponseRedirect(success_url)


# ====================== ORDINI ======================

class OrdineListView(LoginRequiredMixin, ListView):
    """Lista degli ordini con filtri avanzati"""
    model = Ordine
    template_name = 'ordini/ordini/elenco.html'
    context_object_name = 'ordini'
    paginate_by = 25

    def get_queryset(self):
        queryset = Ordine.objects.select_related(
            'prodotto', 'fornitore', 'prodotto__categoria'
        )
        
        # Applica filtri dal form di ricerca
        form = OrdineSearchForm(self.request.GET)
        if form.is_valid():
            q = form.cleaned_data.get('q')
            if q:
                queryset = queryset.filter(
                    Q(numero_ordine__icontains=q) |
                    Q(prodotto__nome_prodotto__icontains=q) |
                    Q(fornitore__nome__icontains=q)
                )
            
            status = form.cleaned_data.get('status')
            if status:
                queryset = queryset.filter(status=status)
            
            fornitore = form.cleaned_data.get('fornitore')
            if fornitore:
                queryset = queryset.filter(fornitore=fornitore)
            
            categoria = form.cleaned_data.get('categoria')
            if categoria:
                queryset = queryset.filter(prodotto__categoria=categoria)
            
            data_da = form.cleaned_data.get('data_da')
            data_a = form.cleaned_data.get('data_a')
            if data_da:
                queryset = queryset.filter(data_creazione_ordine__gte=data_da)
            if data_a:
                queryset = queryset.filter(data_creazione_ordine__lte=data_a)
        
        return queryset.order_by('-data_creazione_ordine')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_form'] = OrdineSearchForm(self.request.GET)
        context['total_ordini'] = self.get_queryset().count()
        
        # Statistiche per status
        context['stats_status'] = {}
        for status_code, status_label in Ordine.StatusOrdine.choices:
            count = self.get_queryset().filter(status=status_code).count()
            context['stats_status'][status_code] = {
                'label': status_label,
                'count': count
            }
        
        return context


class OrdineDetailView(LoginRequiredMixin, OrdineAccessMixin, DetailView):
    """Dettaglio ordine completo"""
    model = Ordine
    template_name = 'ordini/ordini/dettaglio.html'
    context_object_name = 'ordine'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Informazioni ricezione se presente
        if hasattr(self.object, 'ricezione'):
            context['ricezione'] = self.object.ricezione
            context['prodotti_ricevuti'] = self.object.ricezione.prodotti_ricevuti.all()
        
        # Cronologia status (da implementare se necessario)
        context['può_modificare'] = (
            self.request.user.is_staff and 
            self.object.status not in [Ordine.StatusOrdine.COMPLETATO, Ordine.StatusOrdine.ANNULLATO]
        )
        
        # Verifica se in ritardo
        context['in_ritardo'] = self.object.is_in_ritardo()
        
        return context


class OrdineCreateView(LoginRequiredMixin, StaffRequiredMixin, CreateView):
    """Creazione nuovo ordine"""
    model = Ordine
    form_class = OrdineForm
    template_name = 'ordini/ordini/nuovo.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_success_url(self):
        return reverse_lazy('ordini:dettaglio_ordine', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        # Controlla se deve inviare l'ordine immediatamente
        if 'invia_ordine' in self.request.POST:
            form.instance.status = Ordine.StatusOrdine.INVIATO
            form.instance.data_invio_ordine = date.today()
        
        messages.success(self.request, f'Ordine {form.instance.numero_ordine} creato con successo!')
        return super().form_valid(form)


class OrdineUpdateView(LoginRequiredMixin, StaffRequiredMixin, UpdateView):
    """Modifica ordine esistente"""
    model = Ordine
    form_class = OrdineForm
    template_name = 'ordini/ordini/modifica.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_success_url(self):
        return reverse_lazy('ordini:dettaglio_ordine', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        # Permetti modifica solo per bozze
        if self.object.status not in [Ordine.StatusOrdine.BOZZA]:
            messages.error(self.request, "Ordine non modificabile in questo stato")
            return redirect('ordini:dettaglio_ordine', pk=self.object.pk)
        
        messages.success(self.request, f'Ordine {form.instance.numero_ordine} aggiornato con successo!')
        return super().form_valid(form)


class OrdineDeleteView(LoginRequiredMixin, StaffRequiredMixin, DeleteView):
    """Eliminazione ordine (solo bozze)"""
    model = Ordine
    template_name = 'ordini/ordini/elimina.html'
    success_url = reverse_lazy('ordini:elenco_ordini')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        
        # Permetti eliminazione solo per bozze
        if self.object.status != Ordine.StatusOrdine.BOZZA:
            messages.error(request, "È possibile eliminare solo ordini in bozza")
            return redirect('ordini:dettaglio_ordine', pk=self.object.pk)
        
        # Verifica se ha ricezioni associate
        if hasattr(self.object, 'ricezione'):
            messages.error(request, "Impossibile eliminare: ordine ha ricezioni associate")
            return redirect('ordini:dettaglio_ordine', pk=self.object.pk)
        
        success_url = self.get_success_url()
        numero_ordine = self.object.numero_ordine
        self.object.delete()
        messages.success(request, f'Ordine {numero_ordine} eliminato con successo!')
        return HttpResponseRedirect(success_url)


class AggiornaStatoOrdineView(LoginRequiredMixin, StaffRequiredMixin, UpdateView):
    """Aggiornamento stato ordine"""
    model = Ordine
    form_class = AggiornaStatoOrdineForm
    template_name = 'ordini/ordini/aggiorna_stato.html'

    def get_success_url(self):
        return reverse_lazy('ordini:dettaglio_ordine', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        # Log del cambio stato
        old_status = Ordine.objects.get(pk=self.object.pk).status
        new_status = form.cleaned_data['status']
        
        # Invio email automatica se necessario
        if new_status == Ordine.StatusOrdine.INVIATO and not self.object.email_inviata:
            self.invia_email_ordine()
        
        messages.success(
            self.request,
            f'Stato ordine aggiornato da "{old_status}" a "{new_status}"'
        )
        return super().form_valid(form)

    def invia_email_ordine(self):
        """Invia email ordine al fornitore"""
        try:
            # Implementa logica invio email
            self.object.email_inviata = True
            self.object.data_invio_email = timezone.now()
            self.object.save(update_fields=['email_inviata', 'data_invio_email'])
            messages.success(self.request, "Email ordine inviata al fornitore")
        except Exception as e:
            messages.error(self.request, f"Errore invio email: {str(e)}")


# ====================== QUICK ORDER ======================

class QuickOrderView(LoginRequiredMixin, StaffRequiredMixin, FormView):
    """Ordine rapido per prodotti frequenti"""
    form_class = QuickOrderForm
    template_name = 'ordini/ordini/quick_order.html'
    success_url = reverse_lazy('ordini:elenco_ordini')

    def form_valid(self, form):
        # Crea ordine dai dati del form
        ordine = Ordine(
            prodotto=form.cleaned_data['prodotto'],
            fornitore=form.cleaned_data['fornitore'],
            quantita_ordinata=form.cleaned_data['quantita'],
            prezzo_unitario_ordine=form.cleaned_data['prezzo_unitario'],
            data_arrivo_previsto=form.cleaned_data['data_arrivo_previsto'],
            misura=form.cleaned_data['prodotto'].misura,
            creato_da=self.request.user,
            status=Ordine.StatusOrdine.BOZZA
        )
        ordine.save()
        
        messages.success(self.request, f'Ordine rapido {ordine.numero_ordine} creato!')
        return redirect('ordini:dettaglio_ordine', pk=ordine.pk)


# ====================== RICEZIONI ======================

class RicezioneCreateView(LoginRequiredMixin, StaffRequiredMixin, CreateView):
    """Creazione ricezione per ordine"""
    model = Ricezione
    form_class = RicezioneForm
    template_name = 'ordini/ricezioni/nuovo.html'

    def dispatch(self, request, *args, **kwargs):
        self.ordine = get_object_or_404(Ordine, pk=kwargs['ordine_pk'])
        
        # Verifica se esiste già una ricezione
        if hasattr(self.ordine, 'ricezione'):
            messages.info(request, "Ricezione già esistente per questo ordine")
            return redirect('ordini:dettaglio_ricezione', pk=self.ordine.ricezione.pk)
        
        # Verifica stato ordine
        if self.ordine.status not in [
            Ordine.StatusOrdine.INVIATO,
            Ordine.StatusOrdine.CONFERMATO,
            Ordine.StatusOrdine.IN_PRODUZIONE,
            Ordine.StatusOrdine.SPEDITO,
            Ordine.StatusOrdine.IN_TRANSITO
        ]:
            messages.error(request, "L'ordine non è in uno stato ricevibile")
            return redirect('ordini:dettaglio_ordine', pk=self.ordine.pk)
        
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['ordine'] = self.ordine
        return kwargs

    def form_valid(self, form):
        form.instance.ordine = self.ordine
        form.instance.ricevuto_da = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('ordini:dettaglio_ricezione', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['ordine'] = self.ordine
        
        # Formset per prodotti ricevuti
        if self.request.POST:
            context['formset'] = ProdottoRicevutoFormSet(
                self.request.POST,
                instance=self.object if hasattr(self, 'object') else None,
                ordine=self.ordine
            )
        else:
            context['formset'] = ProdottoRicevutoFormSet(
                instance=None,
                ordine=self.ordine,
                initial=[{
                    'prodotto': self.ordine.prodotto,
                    'quantita_ricevuta': self.ordine.quantita_ordinata
                }]
            )
        
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        
        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()
            
            # Crea automaticamente prodotto ricevuto se formset è vuoto
            if not formset.forms[0].cleaned_data:
                ProdottoRicevuto.objects.create(
                    ricezione=self.object,
                    prodotto=self.ordine.prodotto,
                    quantita_ricevuta=self.ordine.quantita_ordinata
                )
            
            messages.success(
                self.request,
                f'Ricezione per ordine {self.ordine.numero_ordine} completata!'
            )
            return super().form_valid(form)
        else:
            return self.form_invalid(form)


class RicezioneDetailView(LoginRequiredMixin, DetailView):
    """Dettaglio ricezione"""
    model = Ricezione
    template_name = 'ordini/ricezioni/dettaglio.html'
    context_object_name = 'ricezione'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['prodotti_ricevuti'] = self.object.prodotti_ricevuti.select_related('prodotto')
        return context


class RicezioneUpdateView(LoginRequiredMixin, StaffRequiredMixin, UpdateView):
    """Modifica ricezione"""
    model = Ricezione
    form_class = RicezioneForm
    template_name = 'ordini/ricezioni/modifica.html'

    def get_success_url(self):
        return reverse_lazy('ordini:dettaglio_ricezione', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, 'Ricezione aggiornata con successo!')
        return super().form_valid(form)


# ====================== MAGAZZINO ======================

class MagazzinoListView(LoginRequiredMixin, ListView):
    """Lista articoli in magazzino con filtri"""
    model = Magazzino
    template_name = 'ordini/magazzino/elenco.html'
    context_object_name = 'magazzino_items'
    paginate_by = 30

    def get_queryset(self):
       queryset = Magazzino.objects.select_related('prodotto', 'prodotto__categoria')
       
       # Applica filtri
       form = MagazzinoFilterForm(self.request.GET)
       if form.is_valid():
           prodotto = form.cleaned_data.get('prodotto')
           if prodotto:
               queryset = queryset.filter(prodotto=prodotto)
           
           categoria = form.cleaned_data.get('categoria')
           if categoria:
               queryset = queryset.filter(prodotto__categoria=categoria)
           
           in_scadenza = form.cleaned_data.get('in_scadenza')
           if in_scadenza:
               queryset = queryset.in_scadenza(giorni=in_scadenza)
           
           scorte_basse = form.cleaned_data.get('scorte_basse')
           if scorte_basse:
               queryset = queryset.scorte_basse()
           
           solo_disponibili = form.cleaned_data.get('solo_disponibili')
           if solo_disponibili:
               queryset = queryset.disponibili()
       
       return queryset.order_by('prodotto__nome_prodotto', 'data_scadenza')

    def get_context_data(self, **kwargs):
       context = super().get_context_data(**kwargs)
       context['filter_form'] = MagazzinoFilterForm(self.request.GET)
       
       # Statistiche rapide
       total_items = self.get_queryset()
       context['total_prodotti'] = total_items.count()
       context['prodotti_disponibili'] = total_items.filter(quantita_in_magazzino__gt=0).count()
       context['prodotti_in_scadenza'] = total_items.in_scadenza().count()
       context['scorte_basse'] = total_items.scorte_basse().count()
       
       return context


class MagazzinoDetailView(LoginRequiredMixin, DetailView):
   """Dettaglio singolo item di magazzino"""
   model = Magazzino
   template_name = 'ordini/magazzino/dettaglio.html'
   context_object_name = 'magazzino_item'

   def get_context_data(self, **kwargs):
       context = super().get_context_data(**kwargs)
       
       # Storico movimenti (se implementato)
       # context['movimenti'] = self.object.movimenti.order_by('-data_movimento')
       
       # Altri lotti dello stesso prodotto
       context['altri_lotti'] = Magazzino.objects.filter(
           prodotto=self.object.prodotto
       ).exclude(pk=self.object.pk).order_by('data_scadenza')
       
       # Ordini recenti del prodotto
       context['ordini_recenti'] = Ordine.objects.filter(
           prodotto=self.object.prodotto
       ).order_by('-data_creazione_ordine')[:5]
       
       return context


class MagazzinoCreateView(LoginRequiredMixin, StaffRequiredMixin, CreateView):
   """Aggiunta manuale prodotto in magazzino"""
   model = Magazzino
   form_class = MagazzinoForm
   template_name = 'ordini/magazzino/nuovo.html'
   success_url = reverse_lazy('ordini:elenco_magazzino')

   def form_valid(self, form):
       messages.success(
           self.request,
           f'Prodotto "{form.instance.prodotto.nome_prodotto}" aggiunto al magazzino!'
       )
       return super().form_valid(form)


class MagazzinoUpdateView(LoginRequiredMixin, StaffRequiredMixin, UpdateView):
   """Modifica item magazzino"""
   model = Magazzino
   form_class = MagazzinoForm
   template_name = 'ordini/magazzino/modifica.html'

   def get_success_url(self):
       return reverse_lazy('ordini:dettaglio_magazzino', kwargs={'pk': self.object.pk})

   def form_valid(self, form):
       messages.success(self.request, 'Dati magazzino aggiornati con successo!')
       return super().form_valid(form)


class MovimentoMagazzinoView(LoginRequiredMixin, StaffRequiredMixin, FormView):
   """Gestione movimento magazzino (carico/scarico)"""
   form_class = MovimentoMagazzinoForm
   template_name = 'ordini/magazzino/movimento.html'

   def dispatch(self, request, *args, **kwargs):
       self.magazzino_item = get_object_or_404(Magazzino, pk=kwargs['pk'])
       return super().dispatch(request, *args, **kwargs)

   def get_form_kwargs(self):
       kwargs = super().get_form_kwargs()
       kwargs['magazzino_item'] = self.magazzino_item
       return kwargs

   def get_context_data(self, **kwargs):
       context = super().get_context_data(**kwargs)
       context['magazzino_item'] = self.magazzino_item
       return context

   def form_valid(self, form):
       tipo_movimento = form.cleaned_data['tipo_movimento']
       quantita = form.cleaned_data['quantita']
       motivo = form.cleaned_data['motivo']
       
       # Applica movimento
       if tipo_movimento == 'carico':
           self.magazzino_item.quantita_in_magazzino += quantita
           azione = 'Caricati'
       elif tipo_movimento == 'scarico':
           self.magazzino_item.quantita_in_magazzino -= quantita
           azione = 'Scaricati'
       else:  # inventario
           self.magazzino_item.quantita_in_magazzino = quantita
           azione = 'Inventario aggiornato a'
       
       self.magazzino_item.save()
       
       # Log movimento (se implementato)
       # MovimentoMagazzino.objects.create(...)
       
       messages.success(
           self.request,
           f'{azione} {quantita} unità per {self.magazzino_item.prodotto.nome_prodotto}'
       )
       
       return redirect('ordini:dettaglio_magazzino', pk=self.magazzino_item.pk)


# ====================== BULK ACTIONS ======================

class BulkActionOrdiniView(LoginRequiredMixin, StaffRequiredMixin, FormView):
   """Azioni bulk per ordini selezionati"""
   form_class = BulkActionForm
   template_name = 'ordini/ordini/bulk_action.html'

   def dispatch(self, request, *args, **kwargs):
       # Ottieni IDs selezionati da POST
       self.selected_ids = request.POST.getlist('selected_ordini')
       if not self.selected_ids:
           messages.error(request, "Nessun ordine selezionato")
           return redirect('ordini:elenco_ordini')
       
       self.selected_ordini = Ordine.objects.filter(id__in=self.selected_ids)
       return super().dispatch(request, *args, **kwargs)

   def get_form_kwargs(self):
       kwargs = super().get_form_kwargs()
       kwargs['selected_count'] = len(self.selected_ids)
       return kwargs

   def get_context_data(self, **kwargs):
       context = super().get_context_data(**kwargs)
       context['selected_ordini'] = self.selected_ordini
       context['selected_count'] = len(self.selected_ids)
       return context

   def form_valid(self, form):
       azione = form.cleaned_data['azione']
       
       if azione == 'cambia_status':
           nuovo_status = form.cleaned_data['nuovo_status']
           count = self.selected_ordini.update(status=nuovo_status)
           messages.success(self.request, f'{count} ordini aggiornati a stato "{nuovo_status}"')
       
       elif azione == 'elimina':
           # Solo bozze
           bozze = self.selected_ordini.filter(status=Ordine.StatusOrdine.BOZZA)
           count = bozze.count()
           bozze.delete()
           if count < len(self.selected_ids):
               messages.warning(
                   self.request,
                   f'Eliminati {count} ordini (solo bozze). Altri ordini non eliminabili.'
               )
           else:
               messages.success(self.request, f'{count} ordini eliminati')
       
       elif azione == 'esporta':
           return self.esporta_ordini_selezionati()
       
       elif azione == 'invia_email':
           return self.invia_email_bulk()
       
       return redirect('ordini:elenco_ordini')

   def esporta_ordini_selezionati(self):
       """Esporta ordini selezionati in CSV"""
       response = HttpResponse(content_type='text/csv')
       response['Content-Disposition'] = 'attachment; filename="ordini_selezionati.csv"'
       
       writer = csv.writer(response)
       writer.writerow([
           'Numero Ordine', 'Prodotto', 'Fornitore', 'Quantità',
           'Prezzo Unitario', 'Totale', 'Status', 'Data Creazione'
       ])
       
       for ordine in self.selected_ordini:
           writer.writerow([
               ordine.numero_ordine,
               ordine.prodotto.nome_prodotto,
               ordine.fornitore.nome,
               ordine.quantita_ordinata,
               ordine.prezzo_unitario_ordine,
               ordine.totale_ordine_ivato,
               ordine.get_status_display(),
               ordine.data_creazione_ordine.strftime('%d/%m/%Y')
           ])
       
       return response

   def invia_email_bulk(self):
       """Invia email per ordini selezionati"""
       count = 0
       for ordine in self.selected_ordini.filter(
           status=Ordine.StatusOrdine.BOZZA,
           email_inviata=False
       ):
           try:
               # Implementa logica invio email
               ordine.email_inviata = True
               ordine.data_invio_email = timezone.now()
               ordine.status = Ordine.StatusOrdine.INVIATO
               ordine.data_invio_ordine = date.today()
               ordine.save()
               count += 1
           except Exception as e:
               continue
       
       messages.success(self.request, f'{count} email inviate con successo')
       return redirect('ordini:elenco_ordini')


# ====================== REPORTS E EXPORT ======================

class ExportOrdiniView(LoginRequiredMixin, StaffRequiredMixin, FormView):
   """Export ordini in vari formati"""
   form_class = ExportOrdiniForm
   template_name = 'ordini/reports/export.html'

   def form_valid(self, form):
       formato = form.cleaned_data['formato']
       data_da = form.cleaned_data.get('data_da')
       data_a = form.cleaned_data.get('data_a')
       status_list = form.cleaned_data.get('status')
       
       # Prepara queryset
       queryset = Ordine.objects.select_related('prodotto', 'fornitore')
       
       if data_da:
           queryset = queryset.filter(data_creazione_ordine__gte=data_da)
       if data_a:
           queryset = queryset.filter(data_creazione_ordine__lte=data_a)
       if status_list:
           queryset = queryset.filter(status__in=status_list)
       
       if formato == 'csv':
           return self.export_csv(queryset)
       elif formato == 'excel':
           return self.export_excel(queryset)
       elif formato == 'pdf':
           return self.export_pdf(queryset)
       
       return redirect('ordini:elenco_ordini')

   def export_csv(self, queryset):
       """Export in formato CSV"""
       response = HttpResponse(content_type='text/csv')
       response['Content-Disposition'] = 'attachment; filename="ordini_export.csv"'
       
       writer = csv.writer(response)
       writer.writerow([
           'Numero Ordine', 'Data Creazione', 'Prodotto', 'EAN', 'Categoria',
           'Fornitore', 'Quantità', 'Misura', 'Prezzo Unitario', 'Sconto %',
           'Totale Netto', 'Totale IVA', 'Status', 'Data Invio', 'Data Arrivo Previsto',
           'Data Ricezione', 'Note Interne', 'Note Fornitore'
       ])
       
       for ordine in queryset:
           writer.writerow([
               ordine.numero_ordine,
               ordine.data_creazione_ordine.strftime('%d/%m/%Y'),
               ordine.prodotto.nome_prodotto,
               ordine.prodotto.ean,
               ordine.prodotto.categoria.nome_categoria,
               ordine.fornitore.nome,
               ordine.quantita_ordinata,
               ordine.get_misura_display(),
               ordine.prezzo_unitario_ordine,
               ordine.sconto_percentuale,
               ordine.prezzo_totale_ordine,
               ordine.totale_ordine_ivato,
               ordine.get_status_display(),
               ordine.data_invio_ordine.strftime('%d/%m/%Y') if ordine.data_invio_ordine else '',
               ordine.data_arrivo_previsto.strftime('%d/%m/%Y') if ordine.data_arrivo_previsto else '',
               ordine.data_ricezione_ordine.strftime('%d/%m/%Y') if ordine.data_ricezione_ordine else '',
               ordine.note_interne,
               ordine.note_fornitore
           ])
       
       return response

   def export_excel(self, queryset):
       """Export in formato Excel (richiede openpyxl)"""
       try:
           import openpyxl
           from openpyxl.styles import Font, PatternFill
       except ImportError:
           messages.error(self.request, "Libreria openpyxl non installata")
           return redirect('ordini:export_ordini')
       
       workbook = openpyxl.Workbook()
       worksheet = workbook.active
       worksheet.title = "Ordini Export"
       
       # Headers
       headers = [
           'Numero Ordine', 'Data Creazione', 'Prodotto', 'Fornitore',
           'Quantità', 'Prezzo Unitario', 'Totale IVA', 'Status'
       ]
       
       # Stile header
       header_font = Font(bold=True)
       header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
       
       for col, header in enumerate(headers, 1):
           cell = worksheet.cell(row=1, column=col)
           cell.value = header
           cell.font = header_font
           cell.fill = header_fill
       
       # Dati
       for row, ordine in enumerate(queryset, 2):
           worksheet.cell(row=row, column=1).value = ordine.numero_ordine
           worksheet.cell(row=row, column=2).value = ordine.data_creazione_ordine
           worksheet.cell(row=row, column=3).value = ordine.prodotto.nome_prodotto
           worksheet.cell(row=row, column=4).value = ordine.fornitore.nome
           worksheet.cell(row=row, column=5).value = ordine.quantita_ordinata
           worksheet.cell(row=row, column=6).value = float(ordine.prezzo_unitario_ordine)
           worksheet.cell(row=row, column=7).value = float(ordine.totale_ordine_ivato)
           worksheet.cell(row=row, column=8).value = ordine.get_status_display()
       
       # Autofit colonne
       for column in worksheet.columns:
           max_length = 0
           column_letter = column[0].column_letter
           for cell in column:
               try:
                   if len(str(cell.value)) > max_length:
                       max_length = len(str(cell.value))
               except:
                   pass
           adjusted_width = min(max_length + 2, 50)
           worksheet.column_dimensions[column_letter].width = adjusted_width
       
       response = HttpResponse(
           content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
       )
       response['Content-Disposition'] = 'attachment; filename="ordini_export.xlsx"'
       workbook.save(response)
       return response

   def export_pdf(self, queryset):
       """Export in formato PDF"""
       from django.template.loader import get_template
       from django.http import HttpResponse
       from weasyprint import HTML, CSS
       from django.template import Context
       
       template = get_template('ordini/reports/ordini_pdf.html')
       
       context = {
           'ordini': queryset,
           'data_export': timezone.now(),
           'total_ordini': queryset.count(),
           'totale_valore': queryset.aggregate(totale=Sum('totale_ordine_ivato'))['totale'] or 0
       }
       
       html_string = template.render(context)
       html = HTML(string=html_string)
       pdf = html.write_pdf()
       
       response = HttpResponse(pdf, content_type='application/pdf')
       response['Content-Disposition'] = 'attachment; filename="ordini_report.pdf"'
       return response


class ReportOrdiniView(LoginRequiredMixin, StaffRequiredMixin, FormView):
   """Generazione report e statistiche ordini"""
   form_class = ReportOrdiniForm
   template_name = 'ordini/reports/report.html'

   def form_valid(self, form):
       tipo_report = form.cleaned_data['tipo_report']
       data_da = form.cleaned_data['data_da']
       data_a = form.cleaned_data['data_a']
       fornitore = form.cleaned_data.get('fornitore')
       categoria = form.cleaned_data.get('categoria')
       
       # Base queryset
       queryset = Ordine.objects.filter(
           data_creazione_ordine__range=[data_da, data_a]
       )
       
       if fornitore:
           queryset = queryset.filter(fornitore=fornitore)
       if categoria:
           queryset = queryset.filter(prodotto__categoria=categoria)
       
       context = {
           'tipo_report': tipo_report,
           'data_da': data_da,
           'data_a': data_a,
           'form': form
       }
       
       if tipo_report == 'ordini_per_fornitore':
           context.update(self.report_ordini_per_fornitore(queryset))
       elif tipo_report == 'prodotti_piu_ordinati':
           context.update(self.report_prodotti_piu_ordinati(queryset))
       elif tipo_report == 'trend_ordini':
           context.update(self.report_trend_ordini(queryset))
       elif tipo_report == 'costi_per_categoria':
           context.update(self.report_costi_per_categoria(queryset))
       elif tipo_report == 'analisi_tempi_consegna':
           context.update(self.report_tempi_consegna(queryset))
       
       return render(self.request, self.template_name, context)

   def report_ordini_per_fornitore(self, queryset):
       """Report ordini raggruppati per fornitore"""
       return {
           'report_data': queryset.values('fornitore__nome').annotate(
               num_ordini=Count('id'),
               totale_speso=Sum('totale_ordine_ivato'),
               media_ordine=Avg('totale_ordine_ivato')
           ).order_by('-totale_speso')
       }

   def report_prodotti_piu_ordinati(self, queryset):
       """Report prodotti più ordinati"""
       return {
           'report_data': queryset.values(
               'prodotto__nome_prodotto',
               'prodotto__categoria__nome_categoria'
           ).annotate(
               quantita_totale=Sum('quantita_ordinata'),
               num_ordini=Count('id'),
               valore_totale=Sum('totale_ordine_ivato')
           ).order_by('-quantita_totale')
       }

   def report_trend_ordini(self, queryset):
       """Report trend ordini nel tempo"""
       from django.db.models import TruncMonth
       
       monthly_data = queryset.annotate(
           month=TruncMonth('data_creazione_ordine')
       ).values('month').annotate(
           num_ordini=Count('id'),
           totale_speso=Sum('totale_ordine_ivato')
       ).order_by('month')
       
       return {
           'report_data': monthly_data,
           'chart_data': {
               'labels': [item['month'].strftime('%Y-%m') for item in monthly_data],
               'ordini': [item['num_ordini'] for item in monthly_data],
               'spesa': [float(item['totale_speso'] or 0) for item in monthly_data]
           }
       }

   def report_costi_per_categoria(self, queryset):
       """Report costi per categoria prodotto"""
       return {
           'report_data': queryset.values(
               'prodotto__categoria__nome_categoria'
           ).annotate(
               num_ordini=Count('id'),
               totale_speso=Sum('totale_ordine_ivato'),
               media_ordine=Avg('totale_ordine_ivato')
           ).order_by('-totale_speso')
       }

   def report_tempi_consegna(self, queryset):
       """Analisi tempi di consegna"""
       ordini_consegnati = queryset.filter(
           data_invio_ordine__isnull=False,
           data_ricezione_ordine__isnull=False
       )
       
       # Calcola tempi di consegna
       tempi_consegna = []
       for ordine in ordini_consegnati:
           giorni_consegna = (ordine.data_ricezione_ordine - ordine.data_invio_ordine).days
           tempi_consegna.append({
               'ordine': ordine,
               'giorni_consegna': giorni_consegna,
               'in_ritardo': ordine.is_in_ritardo()
           })
       
       # Statistiche
       giorni_list = [t['giorni_consegna'] for t in tempi_consegna]
       
       return {
           'report_data': tempi_consegna,
           'statistiche': {
               'tempo_medio': sum(giorni_list) / len(giorni_list) if giorni_list else 0,
               'tempo_minimo': min(giorni_list) if giorni_list else 0,
               'tempo_massimo': max(giorni_list) if giorni_list else 0,
               'ordini_in_ritardo': len([t for t in tempi_consegna if t['in_ritardo']]),
               'percentuale_ritardo': len([t for t in tempi_consegna if t['in_ritardo']]) / len(tempi_consegna) * 100 if tempi_consegna else 0
           }
       }


# ====================== API VIEWS ======================

class ApiSearchProdottiView(LoginRequiredMixin, View):
   """API per ricerca prodotti (per autocomplete)"""
   
   def get(self, request):
       q = request.GET.get('q', '')
       if len(q) < 2:
           return JsonResponse({'results': []})
       
       prodotti = Prodotto.objects.filter(
           Q(nome_prodotto__icontains=q) |
           Q(ean__icontains=q) |
           Q(codice_interno__icontains=q),
           attivo=True
       )[:20]
       
       results = []
       for prodotto in prodotti:
           # Quantità in magazzino
           quantita_magazzino = Magazzino.objects.filter(
               prodotto=prodotto
           ).aggregate(totale=Sum('quantita_in_magazzino'))['totale'] or 0
           
           results.append({
               'id': prodotto.id,
               'text': f"{prodotto.nome_prodotto} ({prodotto.ean})",
               'nome': prodotto.nome_prodotto,
               'ean': prodotto.ean,
               'categoria': prodotto.categoria.nome_categoria,
               'quantita_magazzino': quantita_magazzino,
               'misura': prodotto.get_misura_display(),
               'aliquota_iva': prodotto.aliquota_iva
           })
       
       return JsonResponse({'results': results})


class ApiStatisticheDashboardView(LoginRequiredMixin, View):
   """API per dati dashboard (per aggiornamenti AJAX)"""
   
   def get(self, request):
       # Ordini da ricevere
       ordini_da_ricevere = Ordine.objects.da_ricevere().count()
       
       # Ordini in ritardo
       ordini_in_ritardo = Ordine.objects.in_ritardo().count()
       
       # Prodotti con scorte basse
       scorte_basse = Magazzino.objects.scorte_basse().count()
       
       # Prodotti in scadenza (30 giorni)
       prodotti_in_scadenza = Magazzino.objects.in_scadenza(30).count()
       
       # Valore ordini ultimi 30 giorni
       data_limite = date.today() - timedelta(days=30)
       valore_recente = Ordine.objects.filter(
           data_creazione_ordine__gte=data_limite
       ).aggregate(totale=Sum('totale_ordine_ivato'))['totale'] or 0
       
       return JsonResponse({
           'ordini_da_ricevere': ordini_da_ricevere,
           'ordini_in_ritardo': ordini_in_ritardo,
           'scorte_basse': scorte_basse,
           'prodotti_in_scadenza': prodotti_in_scadenza,
           'valore_ordini_30gg': float(valore_recente),
           'timestamp': timezone.now().isoformat()
       })


class ApiOrdineCalcolaPrezzoView(LoginRequiredMixin, View):
   """API per calcolo automatico prezzi ordine"""
   
   def post(self, request):
       try:
           data = json.loads(request.body)
           prodotto_id = data.get('prodotto_id')
           quantita = int(data.get('quantita', 0))
           prezzo_unitario = Decimal(data.get('prezzo_unitario', '0'))
           sconto_percentuale = Decimal(data.get('sconto_percentuale', '0'))
           pezzi_per_confezione = data.get('pezzi_per_confezione')
           misura = data.get('misura')
           
           if not prodotto_id or quantita <= 0 or prezzo_unitario <= 0:
               return JsonResponse({'error': 'Dati non validi'}, status=400)
           
           prodotto = Prodotto.objects.get(id=prodotto_id)
           
           # Calcola prezzi
           prezzo_scontato = prezzo_unitario * (1 - sconto_percentuale / 100)
           
           if misura == Ordine.Misura.CONFEZIONE and pezzi_per_confezione:
               prezzo_totale = prezzo_scontato * Decimal(pezzi_per_confezione) * quantita
           else:
               prezzo_totale = prezzo_scontato * quantita
           
           # Applica IVA
           aliquota_iva = prodotto.get_aliquota_iva_numerica()
           totale_ivato = prezzo_totale * (Decimal('1.0') + aliquota_iva)
           
           return JsonResponse({
               'prezzo_scontato': float(prezzo_scontato),
               'prezzo_totale': float(prezzo_totale),
               'totale_ivato': float(totale_ivato),
               'aliquota_iva': float(aliquota_iva * 100),
               'risparmio': float(prezzo_unitario - prezzo_scontato) * quantita
           })
           
       except Exception as e:
           return JsonResponse({'error': str(e)}, status=400)


# ====================== UTILITY VIEWS ======================

class ScadenzeOrdiniView(LoginRequiredMixin, TemplateView):
   """Vista dedicate per scadenze e alert"""
   template_name = 'ordini/utility/scadenze.html'

   def get_context_data(self, **kwargs):
       context = super().get_context_data(**kwargs)
       
       # Ordini in ritardo
       context['ordini_in_ritardo'] = Ordine.objects.in_ritardo()
       
       # Ordini da ricevere presto
       data_limite = date.today() + timedelta(days=7)
       context['ordini_arrivo_prossimo'] = Ordine.objects.filter(
           data_arrivo_previsto__lte=data_limite,
           status__in=[
               Ordine.StatusOrdine.INVIATO,
               Ordine.StatusOrdine.CONFERMATO,
               Ordine.StatusOrdine.IN_PRODUZIONE,
               Ordine.StatusOrdine.SPEDITO,
               Ordine.StatusOrdine.IN_TRANSITO
           ]
       )
       
       # Prodotti in scadenza
       context['prodotti_in_scadenza'] = Magazzino.objects.in_scadenza(30)
       
       # Scorte basse
       context['scorte_basse'] = Magazzino.objects.scorte_basse()
       
       return context


class CalcolatoreOrdineView(LoginRequiredMixin, TemplateView):
   """Tool per calcolare quantità ottimali di ordinazione"""
   template_name = 'ordini/utility/calcolatore.html'

   def get_context_data(self, **kwargs):
       context = super().get_context_data(**kwargs)
       
       # Lista prodotti per il calcolatore
       context['prodotti'] = Prodotto.objects.filter(attivo=True).annotate(
           quantita_magazzino=Sum('magazzino__quantita_in_magazzino')
       )
       
       return context